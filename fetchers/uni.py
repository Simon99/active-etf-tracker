"""統一投信 (ezmoney.com.tw) — 主動式 ETF PCF Excel 抓取。

ETF 對照（從 https://www.ezmoney.com.tw/ETF/Transaction/PCF 頁面 scrape 而來）：
  00981A 主動統一台股增長  → fundCode 49YTW
  00403A 主動統一升級50    → fundCode 63YTW
  00988A 主動統一全球創新  → fundCode 61YTW

PCF Excel 結構：
  - sheet 名: '申購買回清單'
  - 前 ~10 行: 基金 metadata（NAV、基金規模、單位數）
  - row 13 (1-indexed): '股票' 標題列
  - row 14: header (股票代號 | 股票名稱 | 股數 | 持股權重)
  - row 15+: 持股資料
"""
from __future__ import annotations

import datetime as dt
import io
import re
import requests
from openpyxl import load_workbook

from .base import FetchResult, UA, to_roc

PCF_EXCEL_URL = 'https://www.ezmoney.com.tw/ETF/Transaction/PCFExcelNPOI'
PCF_PAGE_URL = 'https://www.ezmoney.com.tw/ETF/Transaction/PCF'

REGISTRY = {
    '00981A': (None, '49YTW'),
    '00403A': (None, '63YTW'),
    '00988A': (None, '61YTW'),
}


def fetch(etf_id: str, fund_code: str, date: dt.date) -> FetchResult:
    res = FetchResult(etf_id=etf_id, date=date)
    headers = {'User-Agent': UA, 'Referer': f'{PCF_PAGE_URL}?fundCode={fund_code}'}
    params = {'fundCode': fund_code, 'date': to_roc(date), 'specificDate': 'false'}

    try:
        r = requests.get(PCF_EXCEL_URL, params=params, headers=headers, timeout=30)
    except requests.RequestException as e:
        res.error = f'request failed: {e}'
        return res

    ct = r.headers.get('content-type', '')
    if 'spreadsheet' not in ct:
        # ezmoney 用 HTML error page 回 200 — 抽 data-content 訊息
        m = re.search(r'data-content="([^"]+)"', r.text)
        res.error = f'not an excel: {m.group(1)[:160] if m else ct}'
        return res

    try:
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        ws = wb['申購買回清單']
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        res.error = f'parse excel failed: {e}'
        return res

    # 抽 metadata（前 10 行）
    meta = {}
    for row in rows[:10]:
        if not row or row[0] is None:
            continue
        k = str(row[0]).strip()
        v = row[1] if len(row) > 1 else None
        if v is None:
            continue
        meta[k] = str(v).strip()
    res.meta = meta

    # 找持股表頭 row（'股票代號'）
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == '股票代號':
            header_idx = i
            break
    if header_idx is None:
        res.error = "header row '股票代號' not found"
        return res

    holdings = []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        stock_id = str(row[0]).strip()
        if not re.match(r'^\d{4,6}[A-Z]?$', stock_id):
            continue
        stock_name = (str(row[1]).strip() if row[1] else '')
        shares_raw = row[2]
        weight_raw = row[3] if len(row) > 3 else None

        shares = _parse_int(shares_raw)
        weight = _parse_pct(weight_raw)

        holdings.append({
            'date': date.isoformat(),
            'etf_id': etf_id,
            'stock_id': stock_id,
            'stock_name': stock_name,
            'shares': shares,
            'weight': weight,
            'market_value': None,  # ezmoney PCF 沒提供，後續可用 shares × close 算
        })

    res.holdings = holdings
    return res


def _parse_int(v) -> int | None:
    if v is None:
        return None
    s = str(v).replace(',', '').strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _parse_pct(v) -> float | None:
    if v is None:
        return None
    s = str(v).replace('%', '').replace(',', '').strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# ── self-import: 註冊到 REGISTRY ──
import sys as _sys
_self = _sys.modules[__name__]
REGISTRY = {k: (_self, v[1]) for k, v in REGISTRY.items()}
